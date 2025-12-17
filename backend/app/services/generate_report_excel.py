from openpyxl import Workbook
from io import BytesIO
from datetime import datetime

def generate_excel(data: list[dict], title="Aircraft Report"):
    if not data:
        raise ValueError("No data to generate Excel.")

    # Desired column order
    columns = [
       "registration", "msn", "type", "model", "status",
        "engine_model", "engine_serial_number", "propeller_model", 
        "propeller_serial_number", "base", "engine_arc", "propeller_arc", "created_at"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = title

    # Add header
    ws.append(columns)

    # Add rows
    for row in data:
        cleaned_row = []
        for col in columns:
            value = row.get(col, "")
            if value is None:
                value = ""
            # Format datetime
            if col == "created_at" and value != "":
                value = datetime.fromisoformat(value).strftime("%Y-%m-%d %H:%M:%S")
            # Convert arcs to True/False
            if col in ["engine_arc", "propeller_arc"]:
                value = bool(value)
            cleaned_row.append(value)
        ws.append(cleaned_row)

    # Optional: sort by created_at descending
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    rows.sort(key=lambda x: x[-1], reverse=True)
    # Clear existing rows
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(columns) + 1):
            ws.cell(row=row_idx, column=col_idx).value = None
    # Re-insert sorted rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Save to BytesIO
    file = BytesIO()
    wb.save(file)
    file.seek(0)
    return file
